#!/usr/bin/env python
# coding: utf-8

# In[16]:


import pandas as pd
import numpy as np
import openpyxl 
import os
import sys
import glob
import regex as re
import warnings
from tqdm.auto import tqdm


# # Funzioni

# In[18]:


# 1) FILES NELLA CARTELLA 
def Files_in_dir(path, extension = "xlsx"):
    #set working dir
    os.chdir(path)

    #extension = 'xlsx'

    #crea la lista con i nomi di tutti i file nella cartella 
    all_filenames = [i for i in glob.glob('*.{}'.format(extension))]

    print("Files in Path: " , all_filenames)
    
    return all_filenames


# In[19]:


# 2)  Apertura del workbook Excel
def load_workbook(file):
    
    wb = openpyxl.load_workbook(filename= file , data_only=True)
    print("Loaded File: ", file)
        
    return wb


# In[20]:


# 3) creazione della Lista dei Sheet da usare

def createListaSheets(wb):        
    # eliminare Contents e altri dalla lista delle sheet da eliminare

    
    sheetsCleaned = [x for x in wb.sheetnames if x not in sheetsToDrop]
    
    listaSheets = []
    for sheet in sheetsCleaned:
        if wb[sheet]["A1"].value == "":
            listaSheets.append(sheet)

    return listaSheets


# In[21]:


# 4) drop del template se già presente
def drop_or_add_Template(wb, template):
    #drop del template se già esistente
    if template in wb.sheetnames:
        del wb[template]      
        wb.create_sheet(template)
        print("Template ",template," eliminato e ri-aggiunto")
    
    else:
        wb.create_sheet(template)
        print("Template ",template," aggiunto")
        
    return wb


# In[22]:


# 5) quale template utilizzare
def getTemplateType(wb):
    
    templateType = 1
    for sheet in wb.sheetnames:
        if wb[sheet]["A4"].value == "REQUIREMENTS":
            templateType = 2
    
    print("Template da utilizzare: ", templateType)
    
    return templateType




# 7) creazione template Df
def getTemplateDf(file , paramsDf, listaSheets):
    
    params = ["File", "Sheet"]
    for par in list(paramsDf.index):
        params.append(par)
    #print(params)
    
    templateDf = pd.DataFrame(columns=params)
    templateDf["Sheet"] = listaSheets
    templateDf["File"] = file
    templateDf.set_index("Sheet", inplace = True)
    
    return templateDf


# In[25]:


# 8) FUNZIONE CHE copia il valore della cella 
def catch_value(wb, sheet, paramsDf, parametro, row, col):
    #print("CATCH: - parametro", parametro, " Sheet:", sheet , "ROWCOL",row,col , "\n")
    
    max_row = wb[sheet].max_row
    
    if paramsDf.loc[parametro, "Where"] == "right cell":
        value = wb[sheet].cell(row,col + 1).value

    elif paramsDf.loc[parametro, "Where"] == "left cell":
        value = wb[sheet].cell(row,col -1).value

    elif paramsDf.loc[parametro, "Where"] == "below cell":      
        value = wb[sheet].cell(row+1,col).value

    elif paramsDf.loc[parametro, "Where"] == "same cell":      
        value = wb[sheet].cell(row,col).value  
            
    elif paramsDf.loc[parametro, "Where"] == "extract from cell":
            
        v = wb[sheet].cell(row,col).value  
        patt = paramsDf.loc[parametro, "Pattern"]
            
        if type(v) == str:
            value = re.findall(patt , v)[0]
                    
    #special rule for N eq TEST template2
    elif paramsDf.loc[parametro, "Where"] == "below in fondo":
            
        valori = []
           
        for riga in range(row + 1, max_row + 1):
            value = wb[sheet].cell(riga +1,col).value
            if type(value) == float or type(value)==int:
                valori.append(value) 
            
        if len(valori) > 0:
            value = max(valori)
        else:
            value = ""
            
            
    MissingList = ["MISSING", "#DIV/0!",  "#RIF!" , "#VALORE!"]
    #settare su missing se il valore è nullo o tra i missing
    if pd.isnull(value) == True or value in MissingList:
        value = ""
    
            
       
    #print(value)
    return value


# In[26]:


#9) estrazione dei parametri per un singolo foglio
def parametri_sheet(wb, sheet, templateDf, paramsDf):
    
    min_row = wb[sheet].min_row
    max_row = wb[sheet].max_row

    min_col = wb[sheet].min_column
    max_col = wb[sheet].max_column
   
    for parametro in paramsDf.index:
        
        match = paramsDf.loc[parametro, "Match"]
        m = 0
        
        if type(match) == str:
            
        #print ("Ricerca del parametro: ", parametro, ", tramite match con: ", match)
    
            # iterazione su tutte le celle
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):

                    #valore della cella 
                    cell =  wb[sheet].cell(row,col).value

                    # Match del parametro con la cella
                    if pd.isnull(cell) == False and type(cell)==str and re.match( match, cell):
                        
                        #print(parametro, " trovato in ", wb[sheet].cell(row,col), ":",wb[sheet].cell(row,col).value)
                        m += 1

                        ### Copia nel dataframe "Template"
                        templateDf.loc[sheet, parametro] = catch_value(wb, sheet, paramsDf, parametro, row, col)

                
        if m == 0:
            #print("!!!!! Parametro ", parametro, " non trovato !!!!!")
            warnings.warn("!!!!! ALMENO UN Parametro non trovato !!!!!")

            templateDf.loc[sheet, parametro] = ""
            
        if m > 1:
            print("Trovati più match per parametro ", parametro)
            warnings.warn('Trovati più match per parametro ')
    
    return templateDf


# In[27]:


#Copia dei parametri nel nuovo template
def copy_to_template(wb, templateDf):
    
    # Sistemare le prime colonne
    templateDf.reset_index(inplace=True)
    templateDf.set_index("File", inplace=True)
    templateDf.reset_index(inplace=True)
    
    # copia delle colonne nella prima riga del nuovo template
    letter = 1
    for col in templateDf.columns:
        wb[template].cell(1,letter).value = col
        letter += 1

    # Copia dei valori del dataset nel template
    for riga in range(len(templateDf)):
        for colonna in range(len(templateDf.columns)):
            wb[template].cell(riga + 2, colonna + 1).value =  templateDf.iloc[riga,colonna]
            
    return wb


# In[28]:


def apply_styles(wb, template):
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill
    
    #Bold della prima riga
    BoldFont = Font( bold=True)
    for col in range(wb[template].max_column):
        wb[template].cell(1, col + 1).font = BoldFont
        
    #evidenziare i missing
    '''
    yellow = "00FFFF00"
    MissingFont = Font(color=yellow)
    
    MissingList = ["MISSING", "#DIV/0!",  "#RIF!" ]
    
    for riga in range(wb[template].max_row):
        for colonna in range(wb[template].max_column):
            valore = wb[template].cell(riga+2, colonna +1 ).value
            if valore in MissingList:
                 wb[template].cell(riga+2, colonna +1).fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
    '''             
    return wb


# # Process a single File

# In[29]:


# Tutto il processo 
def create_template_for_file(file , template):
    
    #caricare il wb
    wb = load_workbook(file)
    
    listaSheets = createListaSheets(wb)
    
    # andare avanti solo se ci sono sheet utilizzabili
    if len(listaSheets)>0:
    
        wb = drop_or_add_Template(wb, template)


        print(listaSheets)
        
        # tipo di template e df con la mappa dei parametri
        templateType = getTemplateType(wb)
        paramsDf = getParamsDf(templateType)


        #creazione del dataframe del Template 
        templateDf = getTemplateDf(file , paramsDf, listaSheets)

        #Riempire il templateDf con i parametri
        for sheet in listaSheets:
            print("Inserting Parameters for Sheet: ", sheet)
            templateDf = parametri_sheet(wb,sheet, templateDf, paramsDf)

        #copia del dataset nel Sheet 
        wb = copy_to_template(wb, templateDf)

        #applicare stili
        wb = apply_styles(wb, template)

        #salvare e chiudere il file
        wb.save(file)
        wb.close()
        print("************ File ", file, "saved and closed \n ************")


# # Process all Files in a Directory

# In[32]:


def create_template_directory(path , template = "Parameters Template"):

    all_filenames = Files_in_dir(path)

    for bar in tqdm(range(len(all_filenames)), position=0, leave=True):
        file = all_filenames[bar]
        try:
            create_template_for_file(file, template)
        except:
            print("Something went wrong for file ", file)
            warnings.warn('Problema con almeno un file')

 